"""
Enhanced Excel Logger - Detailed Trade and Hedge Tracking
Tracks every entry/exit of straddle legs and hedges with timestamps
🔥 CLEANED: Removed buffer_target_pct column (PURE SELL hedge strategy)
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os
from config import config


class ExcelLogger:
    """Enhanced logger with detailed trade and hedge tracking"""

    def __init__(self, filename: str = None):
        """Initialize Excel logger with multiple sheets"""
        # Use config file path or default
        self.filename = filename or config.EXCEL_LOG_PATH
        self.wb = None
        self.trade_sheet = None
        self.hedge_sheet = None
        self.leg_sheet = None
        self.trade_row = 2
        self.hedge_row = 2
        self.leg_row = 2
        self.current_trade_id = 0

        self.setup_workbook()

    def setup_workbook(self):
        """Create workbook with three sheets: Trades, Hedges, Legs"""
        try:
            # Backup existing file
            if os.path.exists(self.filename):
                backup_name = f"{self.filename.replace('.xlsx', '')}_backup_{config.get_current_ist_time().strftime('%Y%m%d_%H%M%S')}.xlsx"
                os.rename(self.filename, backup_name)
                print(f"📦 Backup created: {backup_name}")

            # Create new workbook
            self.wb = openpyxl.Workbook()

            # Setup three sheets
            self._setup_trade_sheet()
            self._setup_hedge_sheet()
            self._setup_leg_sheet()

            # Save
            self.wb.save(self.filename)
            print(f"✅ Enhanced Excel initialized: {self.filename}")
            print(f"   📊 3 sheets created: Trades, Hedges, Legs")

        except Exception as e:
            print(f"❌ Error setting up Excel: {str(e)}")

    def _setup_trade_sheet(self):
        """Setup main trade summary sheet"""
        self.trade_sheet = self.wb.active
        self.trade_sheet.title = "Trades"

        headers = [
            "Trade #", "Entry Time", "Exit Time", "Duration (min)",
            "ATM Strike", "Spot Price",
            "CE Entry ₹", "PE Entry ₹", "Total Premium ₹",
            "CE Exit ₹", "PE Exit ₹",
            "CE P&L ₹", "PE P&L ₹",
            "CE Hedge P&L ₹", "PE Hedge P&L ₹", "Total Hedge P&L ₹",
            "Net P&L ₹", "Exit Reason",
            "CE Hedges Used", "PE Hedges Used"
        ]

        self._apply_headers(self.trade_sheet, headers)

    def _setup_hedge_sheet(self):
        """🔥 CLEANED: Setup hedge tracking sheet WITHOUT buffer column"""
        self.hedge_sheet = self.wb.create_sheet("Hedges")

        headers = [
            "Trade #", "Leg Type", "Level", "Strike",
            "Entry Time", "Entry Premium ₹", "Entry Loss %",
            "Exit Time", "Exit Premium ₹", "Exit Loss %",
            "Duration (min)", "Trail Activated",
            "Hedge P&L ₹", "Exit Reason"
        ]

        self._apply_headers(self.hedge_sheet, headers)

    def _setup_leg_sheet(self):
        """Setup detailed leg entry/exit tracking sheet"""
        self.leg_sheet = self.wb.create_sheet("Legs")

        headers = [
            "Trade #", "Leg Type", "Action", "Time", "Strike",
            "Symbol", "Security ID", "Premium ₹",
            "Quantity", "Order Status", "Notes"
        ]

        self._apply_headers(self.leg_sheet, headers)

    def _apply_headers(self, sheet, headers):
        """Apply styled headers to a sheet"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Auto-width columns
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    def log_entry(self, entry_time: datetime, atm_strike: int, ce_premium: float,
                  pe_premium: float, spot_price: float):
        """Log straddle entry"""
        try:
            self.current_trade_id += 1

            # Main trade sheet
            self.trade_sheet.cell(row=self.trade_row, column=1, value=self.current_trade_id)
            self.trade_sheet.cell(row=self.trade_row, column=2, value=entry_time.strftime('%Y-%m-%d %H:%M:%S'))
            self.trade_sheet.cell(row=self.trade_row, column=5, value=atm_strike)
            self.trade_sheet.cell(row=self.trade_row, column=6, value=round(spot_price, 2))
            self.trade_sheet.cell(row=self.trade_row, column=7, value=round(ce_premium, 2))
            self.trade_sheet.cell(row=self.trade_row, column=8, value=round(pe_premium, 2))
            self.trade_sheet.cell(row=self.trade_row, column=9, value=round(ce_premium + pe_premium, 2))

            self.wb.save(self.filename)

            print(f"\n{'=' * 80}")
            print(f"📝 TRADE #{self.current_trade_id} LOGGED TO EXCEL")
            print(f"{'=' * 80}")
            print(f"   Entry Time: {entry_time.strftime('%H:%M:%S')}")
            print(f"   Strike: {atm_strike} | Spot: {spot_price:.2f}")
            print(f"   CE Premium: ₹{ce_premium:.2f} | PE Premium: ₹{pe_premium:.2f}")
            print(f"{'=' * 80}\n")

        except Exception as e:
            print(f"❌ Error logging entry: {str(e)}")

    def log_leg_action(self, leg_type: str, action: str, time: datetime,
                       strike: int, symbol: str, security_id: str,
                       premium: float, quantity: int, order_status: str, notes: str = ""):
        """Log individual leg action (BUY/SELL)"""
        try:
            row = self.leg_row

            self.leg_sheet.cell(row=row, column=1, value=self.current_trade_id)
            self.leg_sheet.cell(row=row, column=2, value=leg_type)
            self.leg_sheet.cell(row=row, column=3, value=action)
            self.leg_sheet.cell(row=row, column=4, value=time.strftime('%Y-%m-%d %H:%M:%S'))
            self.leg_sheet.cell(row=row, column=5, value=strike)
            self.leg_sheet.cell(row=row, column=6, value=symbol)
            self.leg_sheet.cell(row=row, column=7, value=security_id)
            self.leg_sheet.cell(row=row, column=8, value=round(premium, 2))
            self.leg_sheet.cell(row=row, column=9, value=quantity)
            self.leg_sheet.cell(row=row, column=10, value=order_status)
            self.leg_sheet.cell(row=row, column=11, value=notes)

            # Color code by action
            action_cell = self.leg_sheet.cell(row=row, column=3)
            if action == "SELL":
                action_cell.font = Font(color="FF0000", bold=True)  # Red
            elif action == "BUY":
                action_cell.font = Font(color="00FF00", bold=True)  # Green

            self.leg_row += 1
            self.wb.save(self.filename)

            print(f"   📋 Leg Action Logged: {action} {leg_type} @ ₹{premium:.2f} - {order_status}")

        except Exception as e:
            print(f"   ⚠️ Error logging leg action: {str(e)}")

    def log_hedge_entry(self, leg_type: str, level: int, strike: int,
                        entry_time: datetime, entry_premium: float,
                        entry_loss_pct: float, buffer_target_pct: float = 0):
        """
        🔥 CLEANED: Log hedge entry WITHOUT buffer_target_pct
        Note: buffer_target_pct parameter kept for backward compatibility but ignored
        """
        try:
            row = self.hedge_row

            self.hedge_sheet.cell(row=row, column=1, value=self.current_trade_id)
            self.hedge_sheet.cell(row=row, column=2, value=leg_type)
            self.hedge_sheet.cell(row=row, column=3, value=level)
            self.hedge_sheet.cell(row=row, column=4, value=strike)
            self.hedge_sheet.cell(row=row, column=5, value=entry_time.strftime('%Y-%m-%d %H:%M:%S'))
            self.hedge_sheet.cell(row=row, column=6, value=round(entry_premium, 2))
            self.hedge_sheet.cell(row=row, column=7, value=round(entry_loss_pct, 2))
            # 🔥 REMOVED: Column 12 (buffer_target_pct) no longer exists

            self.wb.save(self.filename)

            print(f"\n{'=' * 60}")
            print(f"🛡️ HEDGE L{level} ENTRY LOGGED - Trade #{self.current_trade_id}")
            print(f"{'=' * 60}")
            print(f"   {leg_type} @ Strike {strike}")
            print(f"   Entry Premium: ₹{entry_premium:.2f}")
            print(f"   Entry Loss: {entry_loss_pct:.1f}%")
            print(f"   🔥 PURE SELL HEDGE - No Buffer/Trailing")
            print(f"{'=' * 60}\n")

        except Exception as e:
            print(f"   ⚠️ Error logging hedge entry: {str(e)}")

    def log_hedge_exit(self, leg_type: str, level: int, exit_time: datetime,
                       exit_premium: float, exit_loss_pct: float,
                       trail_activated: bool, hedge_pnl: float, exit_reason: str):
        """Log hedge exit"""
        try:
            # Find the corresponding hedge entry row (work backwards)
            hedge_row_to_update = None
            for row in range(self.hedge_row - 1, 1, -1):
                if (self.hedge_sheet.cell(row=row, column=1).value == self.current_trade_id and
                        self.hedge_sheet.cell(row=row, column=2).value == leg_type and
                        self.hedge_sheet.cell(row=row, column=3).value == level and
                        self.hedge_sheet.cell(row=row, column=8).value is None):  # No exit time yet
                    hedge_row_to_update = row
                    break

            if hedge_row_to_update:
                # Calculate duration
                entry_time_str = self.hedge_sheet.cell(row=hedge_row_to_update, column=5).value
                entry_time = datetime.strptime(entry_time_str, '%Y-%m-%d %H:%M:%S')
                duration = (exit_time - entry_time).total_seconds() / 60

                # Update exit details (columns shifted after buffer removal)
                self.hedge_sheet.cell(row=hedge_row_to_update, column=8, value=exit_time.strftime('%Y-%m-%d %H:%M:%S'))
                self.hedge_sheet.cell(row=hedge_row_to_update, column=9, value=round(exit_premium, 2))
                self.hedge_sheet.cell(row=hedge_row_to_update, column=10, value=round(exit_loss_pct, 2))
                self.hedge_sheet.cell(row=hedge_row_to_update, column=11, value=round(duration, 1))
                self.hedge_sheet.cell(row=hedge_row_to_update, column=12, value="Yes" if trail_activated else "No")

                pnl_cell = self.hedge_sheet.cell(row=hedge_row_to_update, column=13, value=round(hedge_pnl, 2))
                self.hedge_sheet.cell(row=hedge_row_to_update, column=14, value=exit_reason)

                # Color code P&L
                if hedge_pnl > 0:
                    pnl_cell.font = Font(color="00FF00", bold=True)
                elif hedge_pnl < 0:
                    pnl_cell.font = Font(color="FF0000", bold=True)

                self.hedge_row += 1
                self.wb.save(self.filename)

                print(f"\n{'=' * 60}")
                print(f"✅ HEDGE L{level} EXIT LOGGED - Trade #{self.current_trade_id}")
                print(f"{'=' * 60}")
                print(f"   {leg_type} @ Exit Premium: ₹{exit_premium:.2f}")
                print(f"   Duration: {duration:.1f} min")
                print(f"   Trail: {'Yes' if trail_activated else 'No'}")
                print(f"   Hedge P&L: ₹{hedge_pnl:.2f}")
                print(f"   Reason: {exit_reason}")
                print(f"{'=' * 60}\n")

        except Exception as e:
            print(f"   ⚠️ Error logging hedge exit: {str(e)}")

    def log_exit(self, exit_time: datetime, ce_exit_premium: float, pe_exit_premium: float,
                 ce_pnl: float, pe_pnl: float, ce_hedge_pnl: float, pe_hedge_pnl: float,
                 total_pnl: float, exit_reason: str, ce_hedges_used: int = 0, pe_hedges_used: int = 0):
        """Log trade exit with complete P&L breakdown"""
        try:
            # Calculate duration
            entry_time_str = self.trade_sheet.cell(row=self.trade_row, column=2).value
            entry_time = datetime.strptime(entry_time_str, '%Y-%m-%d %H:%M:%S')
            duration = (exit_time - entry_time).total_seconds() / 60

            # Update exit details
            self.trade_sheet.cell(row=self.trade_row, column=3, value=exit_time.strftime('%Y-%m-%d %H:%M:%S'))
            self.trade_sheet.cell(row=self.trade_row, column=4, value=round(duration, 1))
            self.trade_sheet.cell(row=self.trade_row, column=10, value=round(ce_exit_premium, 2))
            self.trade_sheet.cell(row=self.trade_row, column=11, value=round(pe_exit_premium, 2))
            self.trade_sheet.cell(row=self.trade_row, column=12, value=round(ce_pnl, 2))
            self.trade_sheet.cell(row=self.trade_row, column=13, value=round(pe_pnl, 2))
            self.trade_sheet.cell(row=self.trade_row, column=14, value=round(ce_hedge_pnl, 2))
            self.trade_sheet.cell(row=self.trade_row, column=15, value=round(pe_hedge_pnl, 2))
            self.trade_sheet.cell(row=self.trade_row, column=16, value=round(ce_hedge_pnl + pe_hedge_pnl, 2))

            pnl_cell = self.trade_sheet.cell(row=self.trade_row, column=17, value=round(total_pnl, 2))
            self.trade_sheet.cell(row=self.trade_row, column=18, value=exit_reason)
            self.trade_sheet.cell(row=self.trade_row, column=19, value=ce_hedges_used)
            self.trade_sheet.cell(row=self.trade_row, column=20, value=pe_hedges_used)

            # Color code P&L
            if total_pnl > 0:
                pnl_cell.font = Font(color="00FF00", bold=True)
            elif total_pnl < 0:
                pnl_cell.font = Font(color="FF0000", bold=True)

            self.wb.save(self.filename)

            print(f"\n{'=' * 80}")
            print(f"📊 TRADE #{self.current_trade_id} EXIT LOGGED")
            print(f"{'=' * 80}")
            print(f"   Duration: {duration:.1f} minutes")
            print(f"   CE P&L: ₹{ce_pnl:,.2f} | CE Hedges: ₹{ce_hedge_pnl:,.2f}")
            print(f"   PE P&L: ₹{pe_pnl:,.2f} | PE Hedges: ₹{pe_hedge_pnl:,.2f}")
            print(f"   {'─' * 76}")
            print(f"   💰 NET P&L: ₹{total_pnl:,.2f}")
            print(f"   Exit Reason: {exit_reason}")
            print(f"   CE Hedges Used: {ce_hedges_used} | PE Hedges Used: {pe_hedges_used}")
            print(f"{'=' * 80}\n")

            # Move to next row
            self.trade_row += 1

        except Exception as e:
            print(f"❌ Error logging exit: {str(e)}")

    def log_manual_intervention(self, intervention_type: str, leg_type: str,
                                level: int, time: datetime, notes: str = ""):
        """
        🔥 NEW: Log manual hedge interventions
        
        Args:
            intervention_type: 'MANUAL_ADD' or 'MANUAL_EXIT' or 'LEVEL_SKIP' or 'FORCESELL' or 'FORCE_EXIT'
            leg_type: 'CE' or 'PE'
            level: Hedge level
            time: Timestamp
            notes: Additional notes
        """
        try:
            row = self.leg_row

            self.leg_sheet.cell(row=row, column=1, value=self.current_trade_id)
            self.leg_sheet.cell(row=row, column=2, value=leg_type)
            self.leg_sheet.cell(row=row, column=3, value=intervention_type)
            self.leg_sheet.cell(row=row, column=4, value=time.strftime('%Y-%m-%d %H:%M:%S'))
            self.leg_sheet.cell(row=row, column=5, value=0)  # Strike unknown
            self.leg_sheet.cell(row=row, column=6, value=f"Manual L{level}")
            self.leg_sheet.cell(row=row, column=7, value="")  # Security ID
            self.leg_sheet.cell(row=row, column=8, value=0)  # Premium
            self.leg_sheet.cell(row=row, column=9, value=config.LOT_SIZE)
            self.leg_sheet.cell(row=row, column=10, value="MANUAL")
            self.leg_sheet.cell(row=row, column=11, value=notes)

            # Color code manual actions
            action_cell = self.leg_sheet.cell(row=row, column=3)
            action_cell.font = Font(color="FFA500", bold=True)  # Orange

            self.leg_row += 1
            self.wb.save(self.filename)

            print(f"   📋 Manual Intervention Logged: {intervention_type} {leg_type} L{level}")

        except Exception as e:
            print(f"   ⚠️ Error logging manual intervention: {str(e)}")

    def close(self):
        """Close workbook"""
        try:
            if self.wb:
                self.wb.save(self.filename)
                self.wb.close()
                print(f"✅ Excel closed: {self.filename}")
                print(f"   📊 {self.trade_row - 2} trades logged")
                print(f"   🛡️ {self.hedge_row - 2} hedges logged")
                print(f"   📋 {self.leg_row - 2} leg actions logged")
        except Exception as e:
            print(f"❌ Error closing Excel: {str(e)}")
